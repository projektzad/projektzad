# myapp/app/models/group_modify.py
from ldap3 import MODIFY_ADD, MODIFY_DELETE, MODIFY_REPLACE
import openpyxl, csv, json
import re
import uuid
import logging

# Corrected import: Assuming connection_utils.py is in the same directory (models)
# and 'models' directory is in sys.path due to routes.py modification.
# Changed from relative import to direct import.
from connection_utils import create_distinguished_name, domain_to_dn

# Get a logger for this module
logger = logging.getLogger(__name__)


# Function to list all groups, returns dicts with 'cn' and 'dn'
def list_all_groups(conn, domain: str) -> list:
    """
    Retrieves a list of all groups in the specified domain,
    including their common name (cn) and distinguished name (dn).
    Args:
        conn: LDAP connection object.
        domain (str): The LDAP domain to search in (e.g., "testad.local").
    Returns:
        list: A list of dictionaries, where each dict has 'cn' and 'dn' for a group.
    """
    groups_data = []
    # domain_to_dn is now imported from connection_utils
    search_base = domain_to_dn(domain)

    conn.search(search_base, '(objectClass=group)', attributes=['cn', 'distinguishedName'])

    for entry in conn.entries:
        cn_val = None
        dn_val = None

        if hasattr(entry, 'cn') and entry.cn:
            # Handle ldap3.abstract.value.Values object or list of values
            if hasattr(entry.cn, 'value'):
                cn_val_raw = entry.cn.value
                cn_val = str(cn_val_raw[0]) if isinstance(cn_val_raw, list) and cn_val_raw else str(cn_val_raw)
            elif isinstance(entry.cn, list):
                cn_val = str(entry.cn[0]) if entry.cn else None
            else:
                cn_val = str(entry.cn)

        if hasattr(entry, 'distinguishedName') and entry.distinguishedName:
            dn_val = str(entry.distinguishedName.value) if hasattr(entry.distinguishedName, 'value') else str(
                entry.distinguishedName)

        if cn_val and dn_val:
            groups_data.append({
                'cn': cn_val,
                'dn': dn_val
            })
    return groups_data


def add_user_to_group(conn, username: str, users_domain: str, users_ou: str, group: str, group_domain: str,
                      group_ou: str) -> bool:
    """
    Adds a user to a group using their components. Prefer add_user_to_group_by_dn.
    (Note: This function may be deprecated in favor of add_user_to_group_by_dn for clarity.)
    """
    user_dn = create_distinguished_name(username, users_domain, users_ou)
    group_dn = create_distinguished_name(group, group_domain, group_ou, is_group=True)
    return conn.extend.microsoft.add_members_to_groups(user_dn, group_dn)


def remove_user_from_group(conn, username: str, users_domain: str, users_ou: str, group: str, group_domain: str,
                           group_ou: str) -> bool:
    """
    Removes a user from a group using their components. Prefer remove_user_from_group_by_dn.
    (Note: This function may be deprecated in favor of remove_user_from_group_by_dn for clarity.)
    """
    user_dn = create_distinguished_name(username, users_domain, users_ou)
    group_dn = create_distinguished_name(group, group_domain, group_ou, is_group=True)
    return conn.extend.microsoft.remove_members_from_groups(user_dn, group_dn)


def csv_adding_to_groups(conn, file_path: str) -> int:
    processed_count = 0
    try:
        with open(file_path, mode='r', newline='', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            next(reader)  # Skip header
            for row in reader:
                if row and len(row) >= 6:
                    users_canonical_name = row[0].strip()
                    users_domain = row[1].strip()
                    users_ou = row[2].strip()
                    group_canonical_name = row[3].strip()
                    group_domain = row[4].strip()
                    group_ou = row[5].strip()
                    if add_user_to_group(conn, username=users_canonical_name, users_domain=users_domain,
                                         users_ou=users_ou, group=group_canonical_name, group_domain=group_domain,
                                         group_ou=group_ou):
                        processed_count += 1
    except Exception as e:
        logger.error(f"Error processing CSV for adding to groups ({file_path}): {e}", exc_info=True)
    return processed_count


def csv_removing_from_groups(conn, file_path: str) -> int:
    processed_count = 0
    try:
        with open(file_path, mode='r', newline='', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            next(reader)  # Skip header
            for row in reader:
                if row and len(row) >= 6:
                    users_canonical_name = row[0].strip()
                    users_domain = row[1].strip()
                    users_ou = row[2].strip()
                    group_canonical_name = row[3].strip()
                    group_domain = row[4].strip()
                    group_ou = row[5].strip()
                    if remove_user_from_group(conn, username=users_canonical_name, users_domain=users_domain,
                                              users_ou=users_ou, group=group_canonical_name, group_domain=group_domain,
                                              group_ou=group_ou):
                        processed_count += 1
    except Exception as e:
        logger.error(f"Error processing CSV for removing from groups ({file_path}): {e}", exc_info=True)
    return processed_count


def excel_adding_to_groups(conn, file_path: str) -> int:
    processed_count = 0
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        header_skipped = False
        for row in sheet.iter_rows(min_row=1):
            if not header_skipped:
                header_skipped = True
                continue
            row_values = [cell.value for cell in row]
            if row_values and len(row_values) >= 6 and any(row_values):
                users_canonical_name = str(row_values[0]).strip() if row_values[0] is not None else ''
                users_domain = str(row_values[1]).strip() if row_values[1] is not None else ''
                users_ou = str(row_values[2]).strip() if row_values[2] is not None else ''
                group_canonical_name = str(row_values[3]).strip() if row_values[3] is not None else ''
                group_domain = str(row_values[4]).strip() if row_values[4] is not None else ''
                group_ou = str(row_values[5]).strip() if row_values[5] is not None else ''
                if all([users_canonical_name, users_domain, group_canonical_name, group_domain]):
                    if add_user_to_group(conn, username=users_canonical_name, users_domain=users_domain,
                                         users_ou=users_ou, group=group_canonical_name, group_domain=group_domain,
                                         group_ou=group_ou):
                        processed_count += 1
    except Exception as e:
        logger.error(f"Error processing Excel for adding to groups ({file_path}): {e}", exc_info=True)
    return processed_count


def excel_removing_from_groups(conn, file_path: str) -> int:
    processed_count = 0
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        header_skipped = False
        for row in sheet.iter_rows(min_row=1):
            if not header_skipped:
                header_skipped = True
                continue
            row_values = [cell.value for cell in row]
            if row_values and len(row_values) >= 6 and any(row_values):
                users_canonical_name = str(row_values[0]).strip() if row_values[0] is not None else ''
                users_domain = str(row_values[1]).strip() if row_values[1] is not None else ''
                users_ou = str(row_values[2]).strip() if row_values[2] is not None else ''
                group_canonical_name = str(row_values[3]).strip() if row_values[3] is not None else ''
                group_domain = str(row_values[4]).strip() if row_values[4] is not None else ''
                group_ou = str(row_values[5]).strip() if row_values[5] is not None else ''
                if all([users_canonical_name, users_domain, group_canonical_name, group_domain]):
                    if remove_user_from_group(conn, username=users_canonical_name, users_domain=users_domain,
                                              users_ou=users_ou, group=group_canonical_name, group_domain=group_domain,
                                              group_ou=group_ou):
                        processed_count += 1
    except Exception as e:
        logger.error(f"Error processing Excel for removing from groups ({file_path}): {e}", exc_info=True)
    return processed_count


def batch_group_adding(conn, file_path: str) -> int:
    processed_count = 0
    file_extension = file_path.split('.')[-1].lower()
    if file_extension == 'csv':
        processed_count = csv_adding_to_groups(conn, file_path=file_path)
    elif file_extension == 'xlsx':
        processed_count = excel_adding_to_groups(conn, file_path=file_path)
    return processed_count


def batch_group_removing(conn, file_path: str) -> int:
    processed_count = 0
    file_extension = file_path.split('.')[-1].lower()
    if file_extension == 'csv':
        processed_count = csv_removing_from_groups(conn, file_path=file_path)
    elif file_extension == 'xlsx':
        processed_count = excel_removing_from_groups(conn, file_path=file_path)
    return processed_count


def load_json_config(file_path: str) -> dict:
    with open(file_path, 'r') as f:
        data = json.load(f)
    return data


def remove_group(conn, group_cn: str, domain: str, group_container_ou: str) -> bool:
    """
    Deletes a group from Active Directory.
    Args:
        conn: LDAP connection object.
        group_cn (str): The common name (CN) of the group.
        domain (str): The domain of the group (e.g., "testad.local").
        group_container_ou (str): The OU/container string where the group resides (e.g., "CN=Users", "OU=Sales/OU=IT").
    Returns:
        bool: True if deleted, False otherwise.
    """
    try:
        group_dn = create_distinguished_name(username=group_cn, domain=domain, organizational_unit=group_container_ou,
                                             is_group=True)
        logger.info(f"Attempting to delete group with DN: {group_dn}")
        return conn.delete(group_dn)
    except Exception as e:
        logger.error(f"Exception in remove_group for {group_cn} in {group_container_ou}: {e}", exc_info=True)
        if hasattr(conn, 'result'):  # Check if result attribute exists before setting
            conn.result = {'result': -1, 'description': f"Exception during group delete: {e}"}
        return False


def add_new_group(conn, config: dict) -> bool:
    """
    Adds a new group to Active Directory based on a configuration dictionary.
    Args:
        conn: LDAP connection object.
        config (dict): Configuration dictionary containing group attributes.
    Returns:
        bool: True if group added successfully, False otherwise.
    """
    try:
        group_name_cn = config['General']['Group name (pre-Windows 2000)']

        domain_dn_from_connection_info = conn.server.info.other.get('defaultNamingContext', [None])[0]
        if not domain_dn_from_connection_info:
            logger.error("defaultNamingContext not found in LDAP server info.")
            if hasattr(conn, 'result'):
                conn.result = {'result': -1, 'description': 'defaultNamingContext not found'}
            return False

        simple_domain_for_create_dn = domain_dn_from_connection_info.replace('DC=', '').replace(',', '.')

        group_target_ou_string = config['General'].get('Target OU', 'CN=Users')

        group_dn = create_distinguished_name(
            username=group_name_cn,
            domain=simple_domain_for_create_dn,
            organizational_unit=group_target_ou_string,
            is_group=True
        )

        sam_account_name = group_name_cn.replace(' ', '')
        if len(sam_account_name) > 20:
            sam_account_name = sam_account_name[:20]
        if not sam_account_name:
            sam_account_name = str(uuid.uuid4()).replace('-', '')[:20]

        attributes = {
            'objectClass': ['top', 'group'],
            'cn': group_name_cn,
            'sAMAccountName': sam_account_name,
            'groupType': -2147483646  # Global Security Group
        }

        description = config['General'].get('Description')
        if description:
            attributes['description'] = description

        email = config['General'].get('E-mail')
        if email:
            attributes['mail'] = email

        logger.info(f"Attempting to add group with DN: {group_dn} and attributes: {attributes}")
        success = conn.add(group_dn, attributes=attributes)

        if not success:
            logger.error(
                f"LDAP add failed for group {group_name_cn}. DN: {group_dn}. Result: {conn.result if hasattr(conn, 'result') else 'N/A'}")
        return success

    except Exception as e:
        logger.error(
            f"Exception in add_new_group for {config.get('General', {}).get('Group name (pre-Windows 2000)', 'UnknownGroup')}: {e}",
            exc_info=True)
        if hasattr(conn, 'result'):
            conn.result = {'result': -1, 'description': f"Exception during group add: {e}"}
        return False


def process_config_file(conn, file_path: str) -> bool:
    """
    Processes a group configuration JSON file to perform add/remove/modify actions.
    (Note: This function might be deprecated in current routes usage.)
    """
    try:
        config = load_json_config(file_path=file_path)
        action = config.get('action')

        if action == "remove":
            group_dn_from_config = config.get('group_DN')
            if group_dn_from_config:
                return conn.delete(group_dn_from_config)
            logger.warning("process_config_file: Remove action needs group_DN or components in config.")
            return False
        elif action == "add":
            return add_new_group(conn, config)
        elif action == "modify":
            logger.warning("process_config_file: Modify action not implemented.")
            return False  # Not implemented, so return False
        else:
            logger.error(f"process_config_file: Unknown action '{action}' in {file_path}")
            return False
    except Exception as e:
        logger.error(f"Error processing group config file {file_path}: {e}", exc_info=True)
        return False
    # Removed the final 'return True' as it should depend on the action's success


def list_group_members(conn, domain: str, group_dn: str) -> list:
    """
    Retrieves the distinguished names (DNs) of members of a specific group in a given LDAP domain.
    Args:
        conn: LDAP connection object.
        domain (str): The LDAP domain (e.g., "testad.local") used to set the search base.
        group_dn (str): The distinguished name (DN) of the group whose members should be retrieved.
    Returns:
        list: A list of member DNs in the specified group, or an empty list.
    """
    try:
        # Read the entry directly using the group's DN as the search base and a base scope search
        if conn.search(group_dn, '(objectClass=group)', search_scope='BASE',
                       attributes=['member']):  # search_scope='BASE'
            if conn.entries:
                entry = conn.entries[0]
                # entry.member might be a ValuesView object, so access its .values attribute
                if hasattr(entry, 'member') and entry.member and hasattr(entry.member,
                                                                         'values') and entry.member.values:
                    return [str(member_val) for member_val in entry.member.values]
        return []
    except Exception as e:
        logger.error(f"Error listing members of group {group_dn}: {e}", exc_info=True)
        return []


def add_user_to_group_by_dn(conn, user_dn: str, group_dn: str) -> bool:
    """
    Adds a user to a group by their distinguished names (DNs).
    Args:
        conn: LDAP connection object.
        user_dn (str): The distinguished name (DN) of the user.
        group_dn (str): The distinguished name (DN) of the group.
    Returns:
        bool: True if successful, False otherwise.
    """
    try:
        return conn.extend.microsoft.add_members_to_groups(user_dn, group_dn)
    except Exception as e:
        logger.error(f"Failed to add user {user_dn} to group {group_dn}: {e}", exc_info=True)
        if hasattr(conn, 'result'):
            conn.result = {'result': -1, 'description': f"Exception adding member: {e}"}
        return False


def remove_user_from_group_by_dn(conn, user_dn: str, group_dn: str) -> bool:
    """
    Removes a user from a group by their distinguished names (DNs).
    Args:
        conn: LDAP connection object.
        user_dn (str): The distinguished name (DN) of the user.
        group_dn (str): The distinguished name (DN) of the group.
    Returns:
        bool: True if successful, False otherwise.
    """
    try:
        return conn.extend.microsoft.remove_members_from_groups(user_dn, group_dn)
    except Exception as e:
        logger.error(f"Failed to remove user {user_dn} from group {group_dn}: {e}", exc_info=True)
        if hasattr(conn, 'result'):
            conn.result = {'result': -1, 'description': f"Exception removing member: {e}"}
        return False